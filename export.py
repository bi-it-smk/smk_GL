import sqlite3
import pandas as pd
from stats import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

def export_attendance():
    con = sqlite3.connect("data/attendance.db")
    df = pd.read_sql_query(
        "SELECT date, workday, present FROM attendance",
        con
    )
    con.close()
    df["date"] = pd.to_datetime(df["date"])

    year_average = year_avg(df)
    month_average = month_avg(df)
    gen_average = att_avg(df)
    monthly = monthly_averages(df)

    df.to_excel(
        "data/attendance.xlsx",
        index=False
    )
    wb = load_workbook("data/attendance.xlsx")
    ws = wb.active
    ws.title = "Data"
    stats = wb.create_sheet("Statistik")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for cell in ws["A"][1:]:
        cell.number_format = "DD.MM.YYYY"

    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_length = 0

        for cell in ws[letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max_length + 2

    stats["A2"] = "Statistiken"
    stats["B4"] = "Total Average"
    stats["C4"] = gen_average
    stats["C4"].number_format = "0.0%"

    stats["B5"] = "This Year"
    stats["C5"] = year_average
    stats["C5"].number_format = "0.0%"

    stats["B6"] = "This Month"
    stats["C6"] = month_average
    stats["C6"].number_format = "0.0%"

    stats["A2"].font = Font(bold=True, size=20)

    stats["B8"] = "Monat"
    stats["C8"] = "Anwesenheit"

    row = 9
    for month, average in monthly.items():
        stats.cell(row=row, column=2, value=str(month))
        stats.cell(row=row, column=3, value=average)
        stats.cell(row=row, column=3).number_format = "0.0%"
        row += 1

    for column in range(1, stats.max_column + 1):
        letter = get_column_letter(column)
        max_length = 0

        for cell in stats[letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        stats.column_dimensions[letter].width = max_length + 2

    graph = BarChart()
    graph.type = "col"
    graph.style = 10
    graph.title = "Monatliche Anwesendheit"
    graph.y_axis.scaling.min = 0
    graph.y_axis.scaling.max = 1

    data = Reference(
        stats,
        min_col=3,
        min_row=9,
        max_row=stats.max_row
    )

    categories = Reference(
        stats,
        min_col=2,
        min_row=9,
        max_row=stats.max_row
    )
    graph.add_data(data, titles_from_data=True)
    graph.set_categories(categories)
    stats.add_chart(graph, "G7")

    wb.save("data/attendance.xlsx")